import time
from tkinter.filedialog import Open
from numpy import append
import pyautogui
import pyperclip
from PIL import Image, ImageChops
import pytesseract
import cv2
import openpyxl
import numpy

#book = openpyxl.Workbook()
book = openpyxl.Workbook() 
book.create_sheet('lista')
lista = book['lista']

for rows in lista.iter_rows(min_row=2, max_row=3):
    for  cell in rows: 
        lista.append(['Itau'])

book.save('compara2.xlsx')