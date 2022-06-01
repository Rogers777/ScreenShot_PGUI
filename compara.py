import time
from tkinter.filedialog import Open
from turtle import width
from numpy import append
import pyautogui
import pyperclip
from PIL import Image, ImageChops
import pytesseract
import cv2
import openpyxl
import numpy

pyautogui.PAUSE = 1
#abre o windows
pyautogui.press ("win")
pyautogui.PAUSE = 2.5
pyautogui.hotkey ("enter")
pyautogui.PAUSE = 2.5


#copia o link 
#pyperclip.copy('https://www.tiktok.com/search/user?lang=pt-BR&q=itau&t=1651634256929')
pyperclip.copy('https://www.tiktok.com/@itau?lang=pt-BR')


pyautogui.hotkey("ctrl","v")
pyautogui.PAUSE = 4
pyautogui.press("enter")
pyautogui.PAUSE = 7


#printa o local desejado da tela
pyautogui.screenshot('logoItau.png',region=(262.5,192,119,124)) #>>>>> ex region=(x,  y,  hight,  width)

#im2 = pyautogui.screenshot('imagemItau.jpg')


#pyautogui.click(483,269) || clica na posição passada como parametro (comentado)
pyautogui.PAUSE = 2.5

#converte a imagem colorida para um novo arquivo , no caso , para Cinza
img = Image.open('logoItau.png').convert('L')
img.save('logoItau.png')
imagem=cv2.imread("logoItau.png")





#img = Image.open('logoItau.png') #.convert('L')
#img.save('logoItau.png')

#imagem=cv2.imread("logoItau.png")

#img = Image.open('logoItau.png') #.convert('L')
#img.save('logoItau.png')


#Abre a imagem
#img = Image.open('logoItau.png')

#ler a imagem
#imagem=cv2.imread("logoItau.png")

caminho = r"C:\Program Files\Tesseract-OCR"
pytesseract.pytesseract.tesseract_cmd = caminho + r"\\Tesseract.exe"

#chamar biblioteca de xlsx
#book = openpyxl.worksheet()  
#cria tabela e xlsx
#book.create_sheet('lista')

#cria nome da tabela
#lista = book['lista']

#carrega o conteúdo na tabela (nome do arquivo.xlsx)
book = openpyxl.load_workbook('compara2.xlsx') 

#Extrai o texto da imagem
texto = pytesseract.image_to_string(imagem, lang="por")
time.sleep(2)

#laço de repetição for para inserir o texto lido na tabela lista
nome_marca = book ['lista']


for linhas in nome_marca.iter_rows(min_row=2, max_row=3):
    for  cell in linhas: 

        #nome_marca.append([texto]) 
        print([texto])
            

        for linhas in nome_marca.iter_rows(min_row=0, max_row=0): 
                if cell.value == 'Itau':
                    nome_marca.append(['Nome igual'])
        else:
                    nome_marca.append(['Diferente'])
                
                
                
                
book.save('compara2.xlsx')


